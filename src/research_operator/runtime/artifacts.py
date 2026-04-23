from __future__ import annotations

import csv
import json
from zipfile import ZIP_DEFLATED, ZipFile
from pathlib import Path

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

from research_operator.schemas import RunArtifacts, RunQuality, RunResult, RunSummary


def ensure_run_dir(base_dir: Path, run_id: str) -> Path:
    run_dir = base_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


def write_artifacts(result: RunResult, base_dir: Path) -> RunResult:
    run_dir = ensure_run_dir(base_dir, result.run_id)
    manifest_path = run_dir / "run_manifest.json"
    summary_path = run_dir / "run_summary.json"
    report_path = run_dir / "research_report.md"
    findings_path = run_dir / "findings.json"
    quality_path = run_dir / "quality.json"
    html_report_path = run_dir / "research_report.html"
    pdf_report_path = run_dir / "research_report.pdf"
    workbook_path = run_dir / "research_workbook.xlsx"
    bundle_path = run_dir / "delivery_bundle.zip"
    chart_path = run_dir / "source_scores.svg"
    timeline_chart_path = run_dir / "event_timeline.svg"
    source_ledger_path = run_dir / "source_ledger.json"
    entities_path = run_dir / "entities.json"
    entities_csv_path = run_dir / "entities.csv"
    events_path = run_dir / "events.json"
    events_csv_path = run_dir / "events.csv"

    result.artifacts = RunArtifacts(
        manifest_path=manifest_path,
        summary_path=summary_path,
        report_path=report_path,
        findings_path=findings_path,
        quality_path=quality_path,
        html_report_path=html_report_path,
        pdf_report_path=pdf_report_path,
        workbook_path=workbook_path,
        bundle_path=bundle_path,
        chart_path=chart_path,
        timeline_chart_path=timeline_chart_path,
        source_ledger_path=source_ledger_path,
        entities_path=entities_path,
        entities_csv_path=entities_csv_path,
        events_path=events_path,
        events_csv_path=events_csv_path,
    )

    quality = calculate_run_quality(result)
    manifest_path.write_text(
        json.dumps(result.model_dump(mode="json"), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    summary_path.write_text(
        json.dumps(build_run_summary(result, quality).model_dump(mode="json"), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    report_path.write_text(render_markdown_report(result), encoding="utf-8")
    html_report_path.write_text(render_html_report(result), encoding="utf-8")
    write_pdf_report(result, pdf_report_path)
    findings_path.write_text(
        json.dumps([item.model_dump(mode="json") for item in result.findings], indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    quality_path.write_text(
        json.dumps(quality.model_dump(mode="json"), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    source_ledger_path.write_text(
        json.dumps([item.model_dump(mode="json") for item in result.sources], indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    entities_path.write_text(
        json.dumps([item.model_dump(mode="json") for item in result.entities], indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    events_path.write_text(
        json.dumps([item.model_dump(mode="json") for item in result.events], indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    write_csv(
        entities_csv_path,
        ["entity", "category", "source_label", "source_locator"],
        [
            {
                "entity": item.entity,
                "category": item.category,
                "source_label": item.source_label,
                "source_locator": item.source_locator,
            }
            for item in result.entities
        ],
    )
    write_csv(
        events_csv_path,
        ["event_type", "subject", "amount", "event_date", "source_label", "source_locator", "evidence"],
        [
            {
                "event_type": item.event_type,
                "subject": item.subject,
                "amount": item.amount,
                "event_date": item.event_date,
                "source_label": item.source_label,
                "source_locator": item.source_locator,
                "evidence": item.evidence,
            }
            for item in result.events
        ],
    )
    write_workbook(result, workbook_path)
    chart_path.write_text(render_source_score_chart(result), encoding="utf-8")
    timeline_chart_path.write_text(render_event_timeline_chart(result), encoding="utf-8")
    write_delivery_bundle(result, bundle_path)
    return result


def render_markdown_report(result: RunResult) -> str:
    lines = [
        f"# Run {result.run_id}",
        "",
        "## Objective",
        "",
        result.task,
        "",
        "## Executive Summary",
        "",
    ]
    lines.extend(render_executive_summary_lines(result))

    lines.extend([
        "",
        "## Plan",
        "",
    ])
    for step in result.plan.steps:
        lines.append(f"- `{step.id}` {step.title}: {step.description}")

    lines.extend(["", "## Findings", ""])
    for finding in result.findings:
        lines.append(f"- **{finding.title}** ({finding.confidence}): {finding.detail}")

    lines.extend(["", "## Key Evidence", ""])
    lines.extend(render_key_evidence_lines(result))

    lines.extend(["", "## Sources", ""])
    if result.sources:
        for source in result.sources:
            lines.append(
                f"- `{source.kind}` {source.label}: {source.locator} "
                f"(score={source.evidence_score})"
            )
    else:
        lines.append("- No explicit sources attached to this run.")

    lines.extend(["", "## Citations", ""])
    lines.extend(render_citation_lines(result))

    lines.extend(["", "## Limitations", ""])
    lines.extend(render_limitation_lines(result))

    lines.extend(["", "## Structured Outputs", ""])
    lines.append(f"- Summary: `{result.artifacts.summary_path}`")
    lines.append(f"- Entities: `{result.artifacts.entities_path}`")
    lines.append(f"- Entities CSV: `{result.artifacts.entities_csv_path}`")
    lines.append(f"- Events: `{result.artifacts.events_path}`")
    lines.append(f"- Events CSV: `{result.artifacts.events_csv_path}`")
    lines.append(f"- Quality: `{result.artifacts.quality_path}`")
    lines.append(f"- Workbook: `{result.artifacts.workbook_path}`")
    lines.append(f"- PDF Report: `{result.artifacts.pdf_report_path}`")
    lines.append(f"- Delivery Bundle: `{result.artifacts.bundle_path}`")
    lines.append(f"- Source Score Chart: `{result.artifacts.chart_path}`")
    lines.append(f"- Event Timeline Chart: `{result.artifacts.timeline_chart_path}`")

    return "\n".join(lines) + "\n"


def render_html_report(result: RunResult) -> str:
    findings = "".join(
        f"<li><strong>{escape_html(item.title)}</strong> "
        f"<span>({escape_html(item.confidence)})</span>: {escape_html(item.detail)}</li>"
        for item in result.findings
    )
    sources = "".join(
        (
            "<li>"
            f"<strong>{escape_html(source.label)}</strong> "
            f"<span>{escape_html(source.kind)}</span> "
            f"<code>{escape_html(source.locator)}</code>"
            f"<p>{escape_html(source.excerpt)}</p>"
            "</li>"
        )
        for source in result.sources
    )
    if not sources:
        sources = "<li>No explicit sources attached to this run.</li>"

    executive_summary = "".join(
        f"<li>{escape_html(item)}</li>" for item in render_executive_summary_lines(result)
    )
    evidence = "".join(
        f"<li>{escape_html(item)}</li>" for item in render_key_evidence_lines(result)
    )
    citations = "".join(
        f"<li>{escape_html(item)}</li>" for item in render_citation_lines(result)
    )
    limitations = "".join(
        f"<li>{escape_html(item)}</li>" for item in render_limitation_lines(result)
    )

    return f"""<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>DeepResearch Agent Run {escape_html(result.run_id)}</title>
    <style>
      :root {{
        --bg: #f6f2e8;
        --card: #fffdf8;
        --ink: #141312;
        --muted: #665f55;
        --accent: #0f766e;
        --line: #ddd4c5;
      }}
      body {{ font-family: Georgia, 'Iowan Old Style', serif; background: var(--bg); color: var(--ink); margin: 0; }}
      main {{ max-width: 960px; margin: 0 auto; padding: 48px 24px 80px; }}
      h1, h2 {{ margin: 0 0 16px; }}
      .hero {{ padding: 28px; background: linear-gradient(135deg, #fffefb, #efe5d2); border: 1px solid var(--line); border-radius: 20px; }}
      .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 20px; margin-top: 24px; }}
      .card {{ background: var(--card); border: 1px solid var(--line); border-radius: 16px; padding: 20px; }}
      ul {{ padding-left: 20px; }}
      code {{ background: #f1ecdf; padding: 2px 6px; border-radius: 6px; }}
      .meta {{ color: var(--muted); }}
    </style>
  </head>
  <body>
    <main>
      <section class="hero">
        <p class="meta">DeepResearch Agent run</p>
        <h1>{escape_html(result.task)}</h1>
        <p>Run ID: <code>{escape_html(result.run_id)}</code></p>
        <p>Created at: <code>{escape_html(str(result.created_at))}</code></p>
      </section>
      <section class="card" style="margin-top: 24px;">
        <h2>Executive Summary</h2>
        <ul>{executive_summary}</ul>
      </section>
      <section class="grid">
        <article class="card">
          <h2>Plan</h2>
          <ul>
            {"".join(f"<li><strong>{escape_html(step.title)}</strong>: {escape_html(step.description)}</li>" for step in result.plan.steps)}
          </ul>
        </article>
        <article class="card">
          <h2>Findings</h2>
          <ul>{findings}</ul>
        </article>
      </section>
      <section class="grid">
        <article class="card">
          <h2>Entities</h2>
          <p>{len(result.entities)} extracted</p>
        </article>
        <article class="card">
          <h2>Events</h2>
          <p>{len(result.events)} extracted</p>
        </article>
      </section>
      <section class="card" style="margin-top: 24px;">
        <h2>Key Evidence</h2>
        <ul>{evidence}</ul>
      </section>
      <section class="card" style="margin-top: 24px;">
        <h2>Source Score Chart</h2>
        <img alt="Source score chart" src="source_scores.svg" style="max-width: 100%; border-radius: 12px;" />
      </section>
      <section class="card" style="margin-top: 24px;">
        <h2>Event Timeline</h2>
        <img alt="Event timeline chart" src="event_timeline.svg" style="max-width: 100%; border-radius: 12px;" />
      </section>
      <section class="card" style="margin-top: 24px;">
        <h2>Sources</h2>
        <ul>{sources}</ul>
      </section>
      <section class="card" style="margin-top: 24px;">
        <h2>Citations</h2>
        <ul>{citations}</ul>
      </section>
      <section class="card" style="margin-top: 24px;">
        <h2>Limitations</h2>
        <ul>{limitations}</ul>
      </section>
    </main>
  </body>
</html>
"""


def escape_html(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_workbook(result: RunResult, path: Path) -> None:
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["run_id", result.run_id])
    summary_sheet.append(["task", result.task])
    summary_sheet.append(["source_count", len(result.sources)])
    summary_sheet.append(["entity_count", len(result.entities)])
    summary_sheet.append(["event_count", len(result.events)])

    findings_sheet = workbook.create_sheet("findings")
    findings_sheet.append(["title", "detail", "confidence"])
    for item in result.findings:
        findings_sheet.append([item.title, item.detail, item.confidence])

    sources_sheet = workbook.create_sheet("sources")
    sources_sheet.append(["label", "kind", "locator", "excerpt", "content_chars", "provider", "evidence_score"])
    for item in result.sources:
        sources_sheet.append(
            [
                item.label,
                item.kind,
                item.locator,
                item.excerpt,
                item.content_chars,
                item.provider.value,
                item.evidence_score,
            ]
        )

    entities_sheet = workbook.create_sheet("entities")
    entities_sheet.append(["entity", "category", "source_label", "source_locator"])
    for item in result.entities:
        entities_sheet.append([item.entity, item.category, item.source_label, item.source_locator])

    events_sheet = workbook.create_sheet("events")
    events_sheet.append(["event_type", "subject", "amount", "event_date", "source_label", "source_locator", "evidence"])
    for item in result.events:
        events_sheet.append(
            [
                item.event_type,
                item.subject,
                item.amount,
                item.event_date,
                item.source_label,
                item.source_locator,
                item.evidence,
            ]
        )

    workbook.save(path)


def write_delivery_bundle(result: RunResult, path: Path) -> None:
    if result.artifacts is None:
        return
    members = [
        result.artifacts.manifest_path,
        result.artifacts.summary_path,
        result.artifacts.report_path,
        result.artifacts.html_report_path,
        result.artifacts.pdf_report_path,
        result.artifacts.findings_path,
        result.artifacts.quality_path,
        result.artifacts.workbook_path,
        result.artifacts.chart_path,
        result.artifacts.timeline_chart_path,
        result.artifacts.source_ledger_path,
        result.artifacts.entities_path,
        result.artifacts.entities_csv_path,
        result.artifacts.events_path,
        result.artifacts.events_csv_path,
    ]
    with ZipFile(path, "w", compression=ZIP_DEFLATED) as archive:
        for member in members:
            if member.exists():
                archive.write(member, arcname=member.name)


def write_pdf_report(result: RunResult, path: Path) -> None:
    pdf = canvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    margin_x = 48
    y = height - 54

    def new_page() -> None:
        nonlocal y
        pdf.showPage()
        y = height - 54

    def write_line(text: str, font_name: str = "Helvetica", font_size: int = 11) -> None:
        nonlocal y
        if y < 60:
            new_page()
        pdf.setFont(font_name, font_size)
        safe_text = text.encode("latin-1", "replace").decode("latin-1")
        pdf.drawString(margin_x, y, safe_text)
        y -= font_size + 6

    def write_wrapped(text: str, font_name: str = "Helvetica", font_size: int = 11) -> None:
        words = text.split()
        if not words:
            write_line("", font_name, font_size)
            return
        line = ""
        max_width = width - margin_x * 2
        for word in words:
            candidate = word if not line else f"{line} {word}"
            candidate_width = stringWidth(candidate.encode("latin-1", "replace").decode("latin-1"), font_name, font_size)
            if candidate_width <= max_width:
                line = candidate
            else:
                write_line(line, font_name, font_size)
                line = word
        if line:
            write_line(line, font_name, font_size)

    write_line("DeepResearch Agent Report", "Helvetica-Bold", 18)
    write_line(f"Run ID: {result.run_id}", "Helvetica", 10)
    write_line(f"Task: {result.task}", "Helvetica", 10)
    write_line("", "Helvetica", 4)

    write_line("Executive Summary", "Helvetica-Bold", 14)
    for item in render_executive_summary_lines(result):
        write_wrapped(f"- {item}")

    write_line("", "Helvetica", 4)
    write_line("Key Findings", "Helvetica-Bold", 14)
    for item in result.findings[:8]:
        write_wrapped(f"- {item.title} ({item.confidence}): {item.detail}")

    write_line("", "Helvetica", 4)
    write_line("Key Evidence", "Helvetica-Bold", 14)
    for item in render_key_evidence_lines(result):
        write_wrapped(f"- {item}")

    write_line("", "Helvetica", 4)
    write_line("Top Citations", "Helvetica-Bold", 14)
    for item in render_citation_lines(result)[:5]:
        write_wrapped(f"- {item}")

    write_line("", "Helvetica", 4)
    write_line("Structured Outputs", "Helvetica-Bold", 14)
    write_wrapped(f"- Entities CSV: {result.artifacts.entities_csv_path}")
    write_wrapped(f"- Events CSV: {result.artifacts.events_csv_path}")
    write_wrapped(f"- Workbook: {result.artifacts.workbook_path}")
    write_wrapped(f"- Source Chart: {result.artifacts.chart_path}")
    write_wrapped(f"- Timeline Chart: {result.artifacts.timeline_chart_path}")

    pdf.save()


def calculate_run_quality(result: RunResult) -> RunQuality:
    source_count = len(result.sources)
    average_evidence_score = (
        round(sum(item.evidence_score for item in result.sources) / source_count, 3)
        if source_count
        else 0.0
    )
    deliverable_count = len([path for path in deliverable_paths(result) if path.exists()])
    warnings: list[str] = []
    if source_count < 2:
        warnings.append("source_diversity_low")
    if not result.events:
        warnings.append("no_structured_events")
    if average_evidence_score < 0.75:
        warnings.append("evidence_score_low")
    score = min(
        1.0,
        round(
            0.2
            + min(source_count, 4) * 0.12
            + min(len(result.entities), 8) * 0.035
            + min(len(result.events), 5) * 0.06
            + min(deliverable_count, 12) * 0.015
            + min(average_evidence_score, 2.0) * 0.1,
            3,
        ),
    )
    return RunQuality(
        score=score,
        source_count=source_count,
        average_evidence_score=average_evidence_score,
        entity_count=len(result.entities),
        event_count=len(result.events),
        deliverable_count=deliverable_count,
        warnings=warnings,
    )


def build_run_summary(result: RunResult, quality: RunQuality) -> RunSummary:
    if result.artifacts is None:
        primary_deliverables: dict[str, str] = {}
    else:
        primary_deliverables = {
            "bundle": str(result.artifacts.bundle_path),
            "pdf_report": str(result.artifacts.pdf_report_path),
            "html_report": str(result.artifacts.html_report_path),
            "workbook": str(result.artifacts.workbook_path),
            "quality": str(result.artifacts.quality_path),
        }
    return RunSummary(
        run_id=result.run_id,
        task=result.task,
        created_at=result.created_at,
        task_type=result.plan.task_type,
        source_count=len(result.sources),
        finding_count=len(result.findings),
        entity_count=len(result.entities),
        event_count=len(result.events),
        quality_score=quality.score,
        warnings=quality.warnings,
        top_sources=[source.label for source in result.sources[:3]],
        primary_deliverables=primary_deliverables,
    )


def deliverable_paths(result: RunResult) -> list[Path]:
    if result.artifacts is None:
        return []
    return [
        result.artifacts.manifest_path,
        result.artifacts.summary_path,
        result.artifacts.report_path,
        result.artifacts.findings_path,
        result.artifacts.quality_path,
        result.artifacts.html_report_path,
        result.artifacts.pdf_report_path,
        result.artifacts.workbook_path,
        result.artifacts.bundle_path,
        result.artifacts.chart_path,
        result.artifacts.timeline_chart_path,
        result.artifacts.source_ledger_path,
        result.artifacts.entities_path,
        result.artifacts.entities_csv_path,
        result.artifacts.events_path,
        result.artifacts.events_csv_path,
    ]


def render_source_score_chart(result: RunResult) -> str:
    width = 720
    bar_height = 34
    top = 32
    left = 180
    right_pad = 40
    chart_width = width - left - right_pad
    sources = result.sources[:5]
    height = top + max(len(sources), 1) * 52 + 32
    max_score = max((item.evidence_score for item in sources), default=1.0) or 1.0

    bars: list[str] = []
    for index, source in enumerate(sources):
        y = top + index * 52
        bar_width = chart_width * (source.evidence_score / max_score)
        bars.append(
            f'<text x="16" y="{y + 22}" font-size="14" fill="#141312">{escape_html(source.label)}</text>'
        )
        bars.append(
            f'<rect x="{left}" y="{y}" width="{bar_width:.2f}" height="{bar_height}" rx="8" fill="#0f766e" opacity="0.88"/>'
        )
        bars.append(
            f'<text x="{left + bar_width + 12:.2f}" y="{y + 22}" font-size="13" fill="#665f55">{source.evidence_score:.2f}</text>'
        )

    if not sources:
        bars.append('<text x="16" y="60" font-size="14" fill="#665f55">No sources available.</text>')

    return f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}" role="img" aria-label="Source evidence score chart">
  <rect width="{width}" height="{height}" fill="#fffdf8" rx="18"/>
  <text x="16" y="22" font-size="16" font-weight="700" fill="#141312">Source Evidence Scores</text>
  {"".join(bars)}
</svg>
"""


def render_event_timeline_chart(result: RunResult) -> str:
    width = 720
    height = 220
    baseline_y = 112
    left = 80
    right = 660
    events = sorted(
        result.events[:6],
        key=lambda item: (item.event_date or "9999-99-99", item.subject.lower()),
    )

    if not events:
        return f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}" role="img" aria-label="Event timeline chart">
  <rect width="{width}" height="{height}" fill="#fffdf8" rx="18"/>
  <text x="16" y="24" font-size="16" font-weight="700" fill="#141312">Event Timeline</text>
  <text x="16" y="70" font-size="14" fill="#665f55">No dated events were extracted for this run.</text>
</svg>
"""

    step = (right - left) / max(len(events) - 1, 1)
    labels: list[str] = [
        f'<line x1="{left}" y1="{baseline_y}" x2="{right}" y2="{baseline_y}" stroke="#d6cbb6" stroke-width="4" stroke-linecap="round"/>'
    ]
    for index, event in enumerate(events):
        x = left + step * index if len(events) > 1 else (left + right) / 2
        labels.append(
            f'<circle cx="{x:.2f}" cy="{baseline_y}" r="10" fill="#0f766e" opacity="0.92"/>'
        )
        labels.append(
            f'<text x="{x:.2f}" y="{baseline_y - 28}" text-anchor="middle" font-size="12" fill="#665f55">{escape_html(event.event_date or "Undated")}</text>'
        )
        labels.append(
            f'<text x="{x:.2f}" y="{baseline_y + 34}" text-anchor="middle" font-size="13" font-weight="700" fill="#141312">{escape_html(trim_label(event.subject, 22))}</text>'
        )
        labels.append(
            f'<text x="{x:.2f}" y="{baseline_y + 54}" text-anchor="middle" font-size="12" fill="#665f55">{escape_html(trim_label(event.event_type, 18))}</text>'
        )

    return f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}" role="img" aria-label="Event timeline chart">
  <rect width="{width}" height="{height}" fill="#fffdf8" rx="18"/>
  <text x="16" y="24" font-size="16" font-weight="700" fill="#141312">Event Timeline</text>
  {"".join(labels)}
</svg>
"""


def render_executive_summary_lines(result: RunResult) -> list[str]:
    lines = [
        f"This run analyzed {len(result.sources)} source(s) and produced {len(result.findings)} top-level findings.",
        f"Structured extraction yielded {len(result.entities)} entities and {len(result.events)} events.",
    ]
    if result.events:
        first_event = result.events[0]
        lines.append(
            f"Top event signal: {first_event.event_type} involving {first_event.subject} "
            f"{'for ' + first_event.amount if first_event.amount else ''}".strip()
        )
    elif result.sources:
        lines.append(f"Primary evidence source: {result.sources[0].label}.")
    return lines


def render_key_evidence_lines(result: RunResult) -> list[str]:
    if result.events:
        return [
            f"{item.subject} | {item.event_type} | {item.amount or 'n/a'} | {item.event_date or 'n/a'} | {item.source_label}"
            for item in result.events[:5]
        ]
    if result.sources:
        return [
            f"{source.label} | {source.locator} | {source.excerpt}"
            for source in result.sources[:5]
        ]
    return ["No evidence sources attached."]


def render_citation_lines(result: RunResult) -> list[str]:
    if not result.sources:
        return ["No citations available."]
    return [
        f"{source.label} | score={source.evidence_score} | {source.locator} | excerpt: {source.excerpt or 'n/a'}"
        for source in result.sources[:10]
    ]


def render_limitation_lines(result: RunResult) -> list[str]:
    limitations = [
        "This version relies on lightweight extraction and may miss nuanced entity relations.",
        "Confidence is driven by source availability and simple normalization rather than deep verification.",
    ]
    if not result.sources:
        limitations.append("No external evidence sources were attached or discovered for this run.")
    if len(result.sources) < 2:
        limitations.append("Source diversity is limited; conclusions should be treated as preliminary.")
    return limitations


def trim_label(value: str, max_chars: int) -> str:
    if len(value) <= max_chars:
        return value
    return value[: max_chars - 1].rstrip() + "..."
